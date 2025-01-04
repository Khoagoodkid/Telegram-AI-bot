import chromadb
from chromadb.utils import embedding_functions
from openpyxl import load_workbook


columns = [
    "Số Thứ Tự",
    "Họ lót",
    "Tên",
    "Họ và Tên",
    "Số cell",
    "Lớp",
    "Lớp năm trước",
    "Xếp loại",
    "Trung bình năm trước",
    "Giới tính",
    "Ngày sinh",
    "Dân tộc",
    "Tôn giáo",
    "Mã định danh/Căn Cước Công Dân",
    "Đã có Căn Cước Công Dân gắn chip",
    "Số Bảo Hiểm Y Tế",
    "Nơi sinh",
    "Nơi Đăng Ký Khám Chữa Bệnh Ban Đầu",
    "Mã Bệnh Viện",
    "Quê quán",
    "Số Điện Thoại Học Sinh",
    "Email",
    "Hộ khẩu",
    "Hiện tại",
    "Số Điện thoại tin nhắn",
    "Họ tên cha",
    "Năm sinh",
    "Nghề nghiệp",
    "Số ĐT cha",
    "Họ tên Mẹ",
    "Năm sinh",
    "Nghề nghiệp",
    "Số ĐT Mẹ"
]

workbook = load_workbook('utils/docs/syll_12.xlsx')
sheet = workbook.active  # Get the active sheet
text = ""
for i,row in enumerate(sheet.iter_rows(values_only=True)):
    if i >= 2:
        text += " ".join(str(cell)  if cell else '' for i,cell in enumerate(row)) + "\n"
 

COLLECTION_NAME = "syll_12"

client = chromadb.PersistentClient(path="./data")
client.heartbeat()
# client.delete_collection(name=COLLECTION_NAME)
collection = client.create_collection(name= COLLECTION_NAME, embedding_function=embedding_functions.DefaultEmbeddingFunction())
para = text.split('\n')
for i, para in enumerate(para):
    collection.add(documents=[para], ids=[str(i)])




